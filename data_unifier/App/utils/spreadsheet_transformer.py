from openpyxl import load_workbook
import pandas as pd
from pathlib import Path
import re
from app.utils.data_transformer import DataTransformer

class SpreadsheetTransformer:
    def __init__(self, file_path):
        self.file_path = Path(file_path)
        self.wb = load_workbook(self.file_path, data_only=True)
        
    def recognize_core_blocks(self):
        """Extracts core data blocks and contextual metadata from all sheets."""
        all_blocks = []
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            core_block, context = self._extract_core_block(ws)
            if core_block is not None:
                df = pd.DataFrame(core_block['data'], columns=core_block['headers'])
                
                # Attach context columns if found
                for key, value in context.items():
                    df[key] = value
                df['_sheet'] = sheet_name
                
                # Apply transformation for location columns
                df = DataTransformer.transform_wide_to_long(df)
                
                all_blocks.append(df)
                
        if all_blocks:
            return pd.concat(all_blocks, ignore_index=True)
        return pd.DataFrame()  # Empty

    def _extract_core_block(self, ws):
        """
        Heuristic:
        - Find the first row with at least 2 non-empty, non-numeric cells (likely headers)
        - Extract contiguous block below as data until a row of mostly empty cells
        - Extract context from cells above the header row
        """
        rows = list(ws.iter_rows(values_only=True))
        header_row_idx = None
        headers = None
        # 1. Find header row
        for i, row in enumerate(rows):
            non_empty = [cell for cell in row if cell not in [None, ""]]
            if len(non_empty) >= 2 and all(isinstance(cell, str) for cell in non_empty):
                header_row_idx = i
                headers = [str(cell).strip() if cell else "" for cell in row]
                break
        if header_row_idx is None:
            return None, {}

        # 2. Find data block: contiguous rows below header with at least 2 non-empty cells
        data = []
        for row in rows[header_row_idx+1:]:
            if sum(cell not in [None, ""] for cell in row) < 2:
                break
            data.append([cell for cell in row[:len(headers)]])
        
        # 3. Extract context (from rows above header)
        context = {}
        for row in rows[:header_row_idx]:
            for cell in row:
                if isinstance(cell, str):
                    # Simple heuristic: look for "Customer", "Period", etc.
                    m = re.match(r"(Customer|Period|Event)\s*[:\-]?\s*(.+)", cell, re.I)
                    if m:
                        context[m.group(1).strip()] = m.group(2).strip()
        return {'headers': headers, 'data': data}, context
